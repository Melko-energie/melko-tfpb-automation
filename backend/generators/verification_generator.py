"""
Generate verification_donnees.xlsx - Excel report of all validation
anomalies organized by commune with a clean, readable template.
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.config.constants import OUTPUT_DIR
from backend.utils.logger import get_logger

log = get_logger()

# ── Colors ──
GREEN_DARK = "1B5E20"
GREEN_MID = "2E7D32"
GREEN_LIGHT = "C8E6C9"
GREEN_BG = "E8F5E9"
RED_DARK = "B71C1C"
RED_LIGHT = "FFCDD2"
RED_BG = "FFF0F0"
AMBER_DARK = "E65100"
AMBER_LIGHT = "FFE0B2"
AMBER_BG = "FFF8E1"
BLUE_DARK = "1565C0"
BLUE_LIGHT = "BBDEFB"
BLUE_BG = "E3F2FD"
GREY_BG = "F5F5F5"
GREY_BORDER = "E0E0E0"
WHITE = "FFFFFF"

# ── Fonts ──
TITLE_FONT = Font(name="Montserrat", bold=True, size=16, color=GREEN_DARK)
SUBTITLE_FONT = Font(name="Montserrat", bold=True, size=11, color=GREEN_MID)
STAT_LABEL = Font(name="Montserrat", size=10, color="6D4C41")
STAT_VALUE = Font(name="Montserrat", bold=True, size=14)
STAT_ERR = Font(name="Montserrat", bold=True, size=14, color=RED_DARK)
STAT_WARN = Font(name="Montserrat", bold=True, size=14, color=AMBER_DARK)
STAT_INFO = Font(name="Montserrat", bold=True, size=14, color=BLUE_DARK)
STAT_OK = Font(name="Montserrat", bold=True, size=14, color=GREEN_MID)

HEADER_FONT = Font(name="Montserrat", bold=True, size=9, color=WHITE)
HEADER_FILL = PatternFill("solid", fgColor=GREEN_MID)

DATA_FONT = Font(name="Montserrat", size=9)
DATA_FONT_BOLD = Font(name="Montserrat", bold=True, size=9)
MONO_FONT = Font(name="Source Code Pro", size=8)

COMMUNE_OK_FONT = Font(name="Montserrat", bold=True, size=11, color=GREEN_MID)
COMMUNE_WARN_FONT = Font(name="Montserrat", bold=True, size=11, color=AMBER_DARK)
COMMUNE_ERR_FONT = Font(name="Montserrat", bold=True, size=11, color=RED_DARK)
COMMUNE_OK_FILL = PatternFill("solid", fgColor=GREEN_BG)
COMMUNE_WARN_FILL = PatternFill("solid", fgColor=AMBER_BG)
COMMUNE_ERR_FILL = PatternFill("solid", fgColor=RED_BG)

STATUS_OK_FONT = Font(name="Montserrat", bold=True, size=9, color=GREEN_MID)
STATUS_WARN_FONT = Font(name="Montserrat", bold=True, size=9, color=AMBER_DARK)
STATUS_ERR_FONT = Font(name="Montserrat", bold=True, size=9, color=RED_DARK)

ROW_ERR_FILL = PatternFill("solid", fgColor=RED_BG)
ROW_WARN_FILL = PatternFill("solid", fgColor=AMBER_BG)
ROW_INFO_FILL = PatternFill("solid", fgColor=BLUE_BG)
ROW_ALT_FILL = PatternFill("solid", fgColor=GREY_BG)

MONEY_FMT = '#,##0.00 €'
THIN_BORDER = Border(
    left=Side(style="thin", color=GREY_BORDER),
    right=Side(style="thin", color=GREY_BORDER),
    top=Side(style="thin", color=GREY_BORDER),
    bottom=Side(style="thin", color=GREY_BORDER),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
LEFT_CENTER = Alignment(horizontal="left", vertical="center")


def generate_verification(commune_report: dict) -> Path:
    """
    Generate verification_donnees.xlsx from the commune-grouped validation report.
    commune_report comes from validate.group_by_commune().
    """
    log.info("Generation du rapport de verification Excel...")

    wb = Workbook()
    communes = commune_report.get("communes", {})
    global_anomalies = commune_report.get("global_anomalies", [])
    summary = commune_report.get("summary", {})

    # ── Sheet 1: Resume Global ──
    ws1 = wb.active
    ws1.title = "Resume Global"
    ws1.sheet_properties.tabColor = GREEN_MID
    _write_resume(ws1, communes, summary, global_anomalies)

    # ── Sheet 2: Detail par Commune ──
    ws2 = wb.create_sheet("Detail par Commune")
    ws2.sheet_properties.tabColor = "388E3C"
    _write_detail(ws2, communes)

    file_path = OUTPUT_DIR / "verification_donnees.xlsx"
    wb.save(file_path)
    log.info(f"Rapport de verification genere : {file_path}")
    return file_path


def _write_resume(ws, communes, summary, global_anomalies):
    """Write the summary sheet."""
    # Column widths
    widths = {1: 3, 2: 30, 3: 16, 4: 16, 5: 16, 6: 16, 7: 18}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    r = 2

    # Title
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    ws.cell(row=r, column=2, value="Rapport de Verification des Donnees").font = TITLE_FONT
    r += 2

    # Stats row
    total_err = summary.get("ERREUR", 0)
    total_warn = summary.get("ATTENTION", 0)
    total_info = summary.get("INFO", 0)
    total_lignes = summary.get("total_lignes", 0)
    nb_communes = len([c for c in communes if c != "(Commune vide)"])
    nb_ok = len([c for c, s in communes.items() if c != "(Commune vide)" and s["erreurs"] == 0 and s["attention"] == 0])

    stats = [
        ("Lignes analysees", total_lignes, STAT_VALUE),
        ("Communes", nb_communes, STAT_VALUE),
        ("Erreurs", total_err, STAT_ERR),
        ("A verifier", total_warn, STAT_WARN),
        ("Suggestions", total_info, STAT_INFO),
        ("Communes OK", nb_ok, STAT_OK),
    ]

    for i, (label, val, font) in enumerate(stats):
        col = 2 + i
        ws.cell(row=r, column=col, value=label).font = STAT_LABEL
        ws.cell(row=r, column=col).alignment = CENTER
        ws.cell(row=r + 1, column=col, value=val).font = font
        ws.cell(row=r + 1, column=col).alignment = CENTER

    r += 4

    # Commune table
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    ws.cell(row=r, column=2, value="SYNTHESE PAR COMMUNE").font = SUBTITLE_FONT
    r += 1

    headers = ["Commune", "Nb Lignes", "Erreurs", "A verifier", "Suggestions", "Statut"]
    for c, h in enumerate(headers, 2):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 1

    sorted_communes = sorted(
        [(c, s) for c, s in communes.items() if c != "(Commune vide)"],
        key=lambda x: (-(x[1]["erreurs"] * 100 + x[1]["attention"] * 10 + x[1]["info"]))
    )

    for commune, stats in sorted_communes:
        if stats["erreurs"] > 0:
            statut = "A corriger"
            statut_font = STATUS_ERR_FONT
            row_fill = ROW_ERR_FILL
        elif stats["attention"] > 0:
            statut = "A verifier"
            statut_font = STATUS_WARN_FONT
            row_fill = ROW_WARN_FILL
        else:
            statut = "OK"
            statut_font = STATUS_OK_FONT
            row_fill = None

        row_data = [commune, stats["nb_lignes"], stats["erreurs"], stats["attention"], stats["info"], statut]
        for c, val in enumerate(row_data, 2):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = DATA_FONT if c < 7 else statut_font
            cell.border = THIN_BORDER
            cell.alignment = CENTER if c > 2 else LEFT_CENTER
            if row_fill and c != 7:
                cell.fill = row_fill

        r += 1

    # Total row
    r += 1
    ws.cell(row=r, column=2, value="TOTAL").font = DATA_FONT_BOLD
    ws.cell(row=r, column=2).border = THIN_BORDER
    for c, val in [(3, sum(s["nb_lignes"] for _, s in sorted_communes)),
                   (4, total_err), (5, total_warn), (6, total_info)]:
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = DATA_FONT_BOLD
        cell.border = THIN_BORDER
        cell.alignment = CENTER

    # Global anomalies at the bottom
    if global_anomalies:
        r += 3
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        ws.cell(row=r, column=2, value="ANOMALIES GENERALES (hors communes)").font = SUBTITLE_FONT
        r += 1
        for a in global_anomalies:
            ws.cell(row=r, column=2, value=a.get("message", "")).font = DATA_FONT
            ws.cell(row=r, column=3, value=a.get("severite", "")).font = DATA_FONT
            r += 1


def _write_detail(ws, communes):
    """Write the detail sheet with one block per commune."""
    # Column widths
    widths = {1: 3, 2: 10, 3: 13, 4: 16, 5: 14, 6: 55, 7: 25, 8: 25}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    r = 2

    # Title
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.cell(row=r, column=2, value="Detail des Anomalies par Commune").font = TITLE_FONT
    r += 2

    sorted_communes = sorted(
        [(c, s) for c, s in communes.items() if c != "(Commune vide)"],
        key=lambda x: (-(x[1]["erreurs"] * 100 + x[1]["attention"] * 10 + x[1]["info"]))
    )

    for commune, stats in sorted_communes:
        total_issues = stats["erreurs"] + stats["attention"] + stats["info"]

        # ── Commune header block ──
        if stats["erreurs"] > 0:
            c_font, c_fill = COMMUNE_ERR_FONT, COMMUNE_ERR_FILL
        elif stats["attention"] > 0:
            c_font, c_fill = COMMUNE_WARN_FONT, COMMUNE_WARN_FILL
        else:
            c_font, c_fill = COMMUNE_OK_FONT, COMMUNE_OK_FILL

        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        cell = ws.cell(row=r, column=2, value=commune)
        cell.font = c_font
        cell.fill = c_fill
        cell.alignment = LEFT_CENTER
        cell.border = THIN_BORDER
        for col in range(3, 9):
            ws.cell(row=r, column=col).fill = c_fill
            ws.cell(row=r, column=col).border = THIN_BORDER
        r += 1

        # Info line
        types_str = ", ".join(stats.get("types", []))
        info_text = f"{stats['nb_lignes']} lignes  |  Types: {types_str}  |  HT: {stats['total_ht']:,.2f} EUR  |  Degrevement: {stats['total_degrevement']:,.2f} EUR"
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.cell(row=r, column=2, value=info_text).font = Font(name="Montserrat", size=9, color="6D4C41")
        r += 1

        # Counters line
        counters = []
        if stats["erreurs"]:
            counters.append(f"{stats['erreurs']} erreur(s)")
        if stats["attention"]:
            counters.append(f"{stats['attention']} a verifier")
        if stats["info"]:
            counters.append(f"{stats['info']} suggestion(s)")
        if not counters:
            counters.append("Aucun probleme")

        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.cell(row=r, column=2, value="  |  ".join(counters)).font = Font(
            name="Montserrat", bold=True, size=9,
            color=RED_DARK if stats["erreurs"] else (AMBER_DARK if stats["attention"] else GREEN_MID)
        )
        r += 1

        if total_issues == 0:
            r += 2
            continue

        # ── Anomaly table headers ──
        detail_headers = ["Ligne", "Severite", "Categorie", "Colonne", "Message", "Valeur trouvee", "Valeur attendue"]
        for c, h in enumerate(detail_headers, 2):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER
        r += 1

        # ── Anomaly rows, grouped by severity ──
        severity_order = ["ERREUR", "ATTENTION", "INFO"]
        for sev in severity_order:
            items = [a for a in stats["anomalies"] if a["severite"] == sev]
            if not items:
                continue

            for a in items:
                if sev == "ERREUR":
                    row_fill = ROW_ERR_FILL
                    sev_font = Font(name="Montserrat", bold=True, size=9, color=RED_DARK)
                elif sev == "ATTENTION":
                    row_fill = ROW_WARN_FILL
                    sev_font = Font(name="Montserrat", bold=True, size=9, color=AMBER_DARK)
                else:
                    row_fill = ROW_INFO_FILL
                    sev_font = Font(name="Montserrat", bold=True, size=9, color=BLUE_DARK)

                sev_label = {"ERREUR": "Erreur", "ATTENTION": "A verifier", "INFO": "Suggestion"}.get(sev, sev)

                row_data = [
                    a.get("ligne", ""),
                    sev_label,
                    a.get("categorie", ""),
                    a.get("colonne", ""),
                    a.get("message", ""),
                    a.get("valeur", ""),
                    a.get("attendu", ""),
                ]

                for c, val in enumerate(row_data, 2):
                    cell = ws.cell(row=r, column=c, value=val)
                    cell.border = THIN_BORDER
                    cell.fill = row_fill

                    if c == 2:  # Ligne
                        cell.font = MONO_FONT
                        cell.alignment = CENTER
                    elif c == 3:  # Severite
                        cell.font = sev_font
                        cell.alignment = CENTER
                    elif c == 6:  # Message
                        cell.font = DATA_FONT
                        cell.alignment = LEFT_WRAP
                    elif c in (7, 8):  # Valeur/Attendu
                        cell.font = MONO_FONT
                        cell.alignment = LEFT_WRAP
                    else:
                        cell.font = DATA_FONT
                        cell.alignment = LEFT_CENTER

                ws.row_dimensions[r].height = max(30, 15 * (1 + len(str(a.get("message", ""))) // 60))
                r += 1

        # Spacing between communes
        r += 2
