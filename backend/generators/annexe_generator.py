import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from backend.utils.logger import get_logger

log = get_logger()

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=9)
DATA_FONT = Font(name="Arial", size=8)
TOTAL_FILL = PatternFill("solid", fgColor="D6E4F0")
TOTAL_FONT = Font(name="Arial", bold=True, size=9)
CURRENCY_FORMAT = '#,##0.00 €'
PERCENT_FORMAT = '0.00%'
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

ANNEXE_COLUMNS = [
    ("Adresse d'imposition", 15),
    ("Adresse des travaux", 30),
    ("Commune", 12),
    ("Code Postal", 10),
    ("Installateur", 18),
    ("N° Programme", 12),
    ("N°OPERATION", 12),
    ("Numéro de facture", 16),
    ("Date de facture", 12),
    ("TFPB", 6),
    ("Date de paiement", 12),
    ("Classification des travaux", 25),
    ("Nature des travaux", 30),
    ("Détail des travaux retenus", 40),
    ("Montant HT facture ", 14),
    ("Taux de TVA facture", 10),
    ("Retenue de garantie ?", 10),
    ("Taux de retenue", 10),
    ("TVA facture", 12),
    ("Montant TTC facture", 14),
    ("Montant de virement TTC", 14),
    ("Montant de virement HT", 14),
    ("Montant subventions encaisses", 14),
    ("Montant des travaux éligibles retenus H.T", 16),
    ("Montant dégrèvement demandé", 14),
    ("CDIF/ SIP", 15),
    ("Adresse CDIF/SIP", 25),
    ("N° d'avis", 18),
    ("N° Fiscal", 18),
]

MONEY_COLUMNS = {
    "Montant HT facture ", "TVA facture", "Montant TTC facture",
    "Montant de virement TTC", "Montant de virement HT",
    "Montant subventions encaisses",
    "Montant des travaux éligibles retenus H.T",
    "Montant dégrèvement demandé",
}


def generate_annexe(
    segment_df: pd.DataFrame,
    synthesis: dict,
    commune: str,
    work_type: str,
    output_path: Path,
) -> Path:
    """Generate the annexe Excel file for a given segment."""
    log.info(f"Generation annexe Excel : {commune} / {work_type}")

    wb = Workbook()

    # ── Sheet 1: Tableau recapitulatif ──
    ws = wb.active
    ws.title = "Tableau récapitulatif"
    _write_title(ws, commune, work_type)
    _write_data_table(ws, segment_df)
    _write_totals(ws, segment_df, synthesis)

    # ── Sheet 2: Synthese ──
    ws2 = wb.create_sheet("Synthèse")
    _write_synthesis_sheet(ws2, synthesis, commune, work_type)

    output_path.mkdir(parents=True, exist_ok=True)
    file_path = output_path / "annexe.xlsx"
    wb.save(file_path)
    log.info(f"Annexe generee : {file_path}")
    return file_path


def _write_title(ws, commune: str, work_type: str):
    type_labels = {
        "TEE": "TRAVAUX D'ÉCONOMIE D'ÉNERGIE (TEE)",
        "PMR": "TRAVAUX PMR",
        "INDIS_TEE": "TRAVAUX INDISSOCIABLEMENT LIÉS (TEE)",
        "INDIS_PMR": "TRAVAUX INDISSOCIABLEMENT LIÉS (PMR)",
    }
    type_label = type_labels.get(work_type, work_type)
    ws.merge_cells("A1:AC1")
    title_cell = ws["A1"]
    title_cell.value = f"TABLEAU RÉCAPITULATIF DE DEMANDE DE REMBOURSEMENT TFPB 2024 - {commune.upper()} - {type_label}"
    title_cell.font = Font(name="Arial", bold=True, size=11, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30


def _write_data_table(ws, df: pd.DataFrame):
    start_row = 3

    # Headers
    for col_idx, (col_name, width) in enumerate(ANNEXE_COLUMNS, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[start_row].height = 40

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), start_row + 1):
        for col_idx, (col_name, _) in enumerate(ANNEXE_COLUMNS, 1):
            value = row.get(col_name, "")
            if pd.isna(value):
                value = ""
            elif col_name in MONEY_COLUMNS and isinstance(value, (int, float)):
                pass  # keep numeric
            elif hasattr(value, "strftime"):
                value = value.strftime("%d/%m/%Y")

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if col_name in MONEY_COLUMNS and isinstance(value, (int, float)):
                cell.number_format = CURRENCY_FORMAT
            elif col_name == "Taux de TVA facture" and isinstance(value, (int, float)):
                cell.number_format = PERCENT_FORMAT

        ws.row_dimensions[row_idx].height = 60


def _write_totals(ws, df: pd.DataFrame, synthesis: dict):
    last_data_row = ws.max_row
    total_row = last_data_row + 2

    # Total TVA 5.5%
    tva_55 = synthesis["tva_groups"].get(0.055, {}).get("montant_ht", 0)
    tva_10 = synthesis["tva_groups"].get(0.1, {}).get("montant_ht", 0)
    tva_20 = synthesis["tva_groups"].get(0.2, {}).get("montant_ht", 0)

    totals_data = [
        ("Total travaux avec taux TVA 5,5% appliqué", tva_55),
        ("Total travaux avec taux TVA 10%", tva_10),
        ("Total travaux avec taux TVA 20%", tva_20),
        ("", ""),
        ("Totaux paiements factures", synthesis["total_ht_eligible"]),
        ("Montant des subventions encaissées en 2023", synthesis["total_subventions"]),
        ("Montant des travaux éligibles retenus", synthesis["base_nette"]),
        ("Montant de dégrèvement demandé", synthesis["total_degrevement"]),
    ]

    for i, (label, value) in enumerate(totals_data):
        r = total_row + i
        label_cell = ws.cell(row=r, column=23, value=label)
        label_cell.font = TOTAL_FONT
        label_cell.fill = TOTAL_FILL
        label_cell.border = THIN_BORDER

        if value != "":
            val_cell = ws.cell(row=r, column=24, value=value)
            val_cell.font = TOTAL_FONT
            val_cell.fill = TOTAL_FILL
            val_cell.number_format = CURRENCY_FORMAT
            val_cell.border = THIN_BORDER


def _write_synthesis_sheet(ws, synthesis: dict, commune: str, work_type: str):
    synth_labels = {
        "TEE": "Économies d'énergie (TEE)",
        "PMR": "Travaux PMR",
        "INDIS_TEE": "Travaux indissociablement liés (TEE)",
        "INDIS_PMR": "Travaux indissociablement liés (PMR)",
    }
    type_label = synth_labels.get(work_type, work_type)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20

    data = [
        ("SYNTHÈSE", ""),
        ("", ""),
        ("Commune", commune),
        ("Type de travaux", type_label),
        ("", ""),
        ("Nombre d'opérations", synthesis["nb_operations"]),
        ("Nombre de factures", synthesis["nb_factures"]),
        ("", ""),
        ("Montant total éligible HT", synthesis["total_ht_eligible"]),
        ("Subventions encaissées", synthesis["total_subventions"]),
        ("Base nette de dégrèvement", synthesis["base_nette"]),
        ("Montant de dégrèvement demandé (25%)", synthesis["total_degrevement"]),
        ("", ""),
        ("DÉTAIL PAR TAUX DE TVA", ""),
    ]

    for rate, info in sorted(synthesis["tva_groups"].items()):
        pct = f"{rate*100:.1f}%"
        data.append((f"  TVA {pct}", info["montant_ht"]))

    data.append(("", ""))
    data.append(("DÉTAIL PAR OPÉRATION", ""))

    for op_id, op_info in synthesis["operations"].items():
        data.append((f"  Opération {op_id}", op_info["montant_ht_total"]))

    for row_idx, (label, value) in enumerate(data, 1):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        val_cell = ws.cell(row=row_idx, column=2, value=value)

        if row_idx == 1:
            label_cell.font = Font(name="Arial", bold=True, size=14, color="1F4E79")
        elif label in ("DÉTAIL PAR TAUX DE TVA", "DÉTAIL PAR OPÉRATION"):
            label_cell.font = Font(name="Arial", bold=True, size=10, color="1F4E79")
        elif label and not label.startswith("  "):
            label_cell.font = Font(name="Arial", bold=True, size=10)
        else:
            label_cell.font = DATA_FONT

        if isinstance(value, (int, float)) and value != 0:
            val_cell.number_format = CURRENCY_FORMAT
            val_cell.font = Font(name="Arial", size=10)
