"""
Generate a recap Excel synthesis file after processing.
Placed in /output/recap_synthese.xlsx
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from backend.config.constants import OUTPUT_DIR
from backend.utils.logger import get_logger

log = get_logger()

# ── Styles ──
TITLE_FONT = Font(name="Montserrat", bold=True, size=14, color="1B5E20")
SECTION_FONT = Font(name="Montserrat", bold=True, size=11, color="2E7D32")
HEADER_FONT = Font(name="Montserrat", bold=True, size=9, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2E7D32")
DATA_FONT = Font(name="Montserrat", size=9)
DATA_FONT_BOLD = Font(name="Montserrat", bold=True, size=9)
TOTAL_FILL = PatternFill("solid", fgColor="C8E6C9")
TOTAL_FONT = Font(name="Montserrat", bold=True, size=10, color="1B5E20")
WARN_FILL = PatternFill("solid", fgColor="FFF3E0")
OK_FONT = Font(name="Montserrat", bold=True, size=9, color="2E7D32")
WARN_FONT = Font(name="Montserrat", bold=True, size=9, color="E65100")
MONEY_FMT = '#,##0.00 €'
PCT_FMT = '0.00%'
THIN_BORDER = Border(
    left=Side(style="thin", color="C8E6C9"),
    right=Side(style="thin", color="C8E6C9"),
    top=Side(style="thin", color="C8E6C9"),
    bottom=Side(style="thin", color="C8E6C9"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def generate_recap(segments_results: list) -> Path:
    """
    Generate recap_synthese.xlsx from the list of segment results.
    Each segment result is a dict with keys:
        commune, type, nb_lignes, total_degrevement, total_ht,
        nb_operations, synthesis (full synthesis dict)
    """
    log.info("Generation du recap de synthese...")
    wb = Workbook()

    # ── Sheet 1: Synthese globale ──
    ws = wb.active
    ws.title = "Synthese TFPB"
    _write_synthese_sheet(ws, segments_results)

    # ── Sheet 2: Detail par commune ──
    ws2 = wb.create_sheet("Detail par commune")
    _write_detail_sheet(ws2, segments_results)

    # ── Sheet 3: Detail TVA ──
    ws3 = wb.create_sheet("Detail TVA")
    _write_tva_sheet(ws3, segments_results)

    file_path = OUTPUT_DIR / "recap_synthese.xlsx"
    wb.save(file_path)
    log.info(f"Recap genere : {file_path}")
    return file_path


def _write_synthese_sheet(ws, results):
    ws.sheet_properties.tabColor = "2E7D32"

    # Column widths
    for col, w in [(1, 5), (2, 40), (3, 22), (4, 22), (5, 15)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    r = 2
    ws.cell(row=r, column=2, value="Synthese TFPB - Recapitulatif General").font = TITLE_FONT
    r += 2

    # ── Totaux globaux ──
    ws.cell(row=r, column=2, value="TOTAUX GLOBAUX").font = SECTION_FONT
    r += 1
    headers = ["Indicateur", "Montant"]
    for c, h in enumerate(headers, 2):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    r += 1

    total_ht = sum(s.get("total_ht", 0) for s in results)
    total_deg = sum(s.get("total_degrevement", 0) for s in results)
    total_sub = sum(s.get("synthesis", {}).get("total_subventions", 0) for s in results)
    total_base = total_ht - total_sub

    globals_data = [
        ("Nombre de communes", len(set(s["commune"] for s in results))),
        ("Nombre de dossiers (commune/type)", len(results)),
        ("Nombre total de lignes", sum(s.get("nb_lignes", 0) for s in results)),
        ("Total HT Eligible", total_ht),
        ("Total Subventions", total_sub),
        ("Base nette de degrevement", total_base),
        ("Degrevement Total (25%)", total_deg),
        ("Taux de degrevement", 0.25),
    ]

    for label, val in globals_data:
        ws.cell(row=r, column=2, value=label).font = DATA_FONT_BOLD
        ws.cell(row=r, column=2).border = THIN_BORDER
        c3 = ws.cell(row=r, column=3, value=val)
        c3.border = THIN_BORDER
        c3.alignment = CENTER
        if isinstance(val, float) and val < 1:
            c3.number_format = PCT_FMT
            c3.font = DATA_FONT_BOLD
        elif isinstance(val, (int, float)) and val > 1:
            c3.number_format = MONEY_FMT if "Nombre" not in label else '0'
            c3.font = DATA_FONT
        r += 1

    # Highlight degrevement total
    ws.cell(row=r - 2, column=2).fill = TOTAL_FILL
    ws.cell(row=r - 2, column=2).font = TOTAL_FONT
    ws.cell(row=r - 2, column=3).fill = TOTAL_FILL
    ws.cell(row=r - 2, column=3).font = TOTAL_FONT

    r += 2

    # ── Detail par commune (sorted by degrevement desc) ──
    ws.cell(row=r, column=2, value="DETAIL PAR COMMUNE").font = SECTION_FONT
    r += 1

    detail_headers = ["Commune", "Type", "Nb lignes", "Nb operations", "Total HT Eligible",
                       "Subventions", "Base nette", "Degrevement", "Taux",
                       "Degrev. calcule (25%)", "Ecart", "Statut"]
    for c, h in enumerate(detail_headers, 2):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.column_dimensions[get_column_letter(2)].width = 25
    for c in range(3, 14):
        ws.column_dimensions[get_column_letter(c)].width = 18
    r += 1

    sorted_results = sorted(results, key=lambda x: x.get("total_degrevement", 0), reverse=True)

    for seg in sorted_results:
        synth = seg.get("synthesis", {})
        ht = seg.get("total_ht", 0)
        sub = synth.get("total_subventions", 0)
        base = ht - sub
        deg = seg.get("total_degrevement", 0)
        deg_calcule = base * 0.25
        ecart = abs(deg - deg_calcule)
        statut = "OK" if ecart < 1 else "A verifier"

        row_data = [
            seg.get("commune", ""),
            seg.get("type", ""),
            seg.get("nb_lignes", 0),
            seg.get("nb_operations", 0),
            ht,
            sub,
            base,
            deg,
            0.25,
            deg_calcule,
            ecart,
            statut,
        ]

        for c, val in enumerate(row_data, 2):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = THIN_BORDER
            cell.font = DATA_FONT

            if c in (6, 7, 8, 9, 11, 12) and isinstance(val, (int, float)):
                cell.number_format = MONEY_FMT
                cell.alignment = CENTER
            elif c == 10:
                cell.number_format = PCT_FMT
                cell.alignment = CENTER
            elif c == 4:
                cell.alignment = CENTER
            elif c == 5:
                cell.alignment = CENTER
            elif c == 13:
                cell.number_format = MONEY_FMT
                cell.alignment = CENTER

            # Statut column
            if c == 13:
                if val == "OK":
                    cell.font = OK_FONT
                    cell.number_format = '@'
                else:
                    cell.font = WARN_FONT
                    cell.fill = WARN_FILL
                    cell.number_format = '@'

        r += 1

    # Total row
    ws.cell(row=r, column=2, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=2).fill = TOTAL_FILL
    ws.cell(row=r, column=2).border = THIN_BORDER
    for c, val in [(4, sum(s.get("nb_lignes", 0) for s in results)),
                   (6, total_ht), (7, total_sub), (8, total_base),
                   (9, total_deg), (11, total_base * 0.25)]:
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        cell.number_format = MONEY_FMT if c != 4 else '0'
        cell.alignment = CENTER


def _write_detail_sheet(ws, results):
    """One row per commune (merged TEE+PMR)."""
    ws.sheet_properties.tabColor = "388E3C"

    headers = ["Commune", "TEE - HT Eligible", "TEE - Degrevement",
               "PMR - HT Eligible", "PMR - Degrevement",
               "Total HT", "Total Degrevement"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    for c in range(1, 8):
        ws.column_dimensions[get_column_letter(c)].width = 22

    # Group by commune
    communes = {}
    for seg in results:
        c = seg.get("commune", "")
        if c not in communes:
            communes[c] = {"TEE_ht": 0, "TEE_deg": 0, "PMR_ht": 0, "PMR_deg": 0}
        if seg.get("type") == "TEE":
            communes[c]["TEE_ht"] += seg.get("total_ht", 0)
            communes[c]["TEE_deg"] += seg.get("total_degrevement", 0)
        else:
            communes[c]["PMR_ht"] += seg.get("total_ht", 0)
            communes[c]["PMR_deg"] += seg.get("total_degrevement", 0)

    r = 2
    for commune, vals in sorted(communes.items(), key=lambda x: -(x[1]["TEE_deg"] + x[1]["PMR_deg"])):
        total_ht = vals["TEE_ht"] + vals["PMR_ht"]
        total_deg = vals["TEE_deg"] + vals["PMR_deg"]
        row = [commune, vals["TEE_ht"], vals["TEE_deg"], vals["PMR_ht"], vals["PMR_deg"], total_ht, total_deg]
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if c > 1:
                cell.number_format = MONEY_FMT
                cell.alignment = CENTER
        r += 1

    # Totals
    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=1).fill = TOTAL_FILL
    ws.cell(row=r, column=1).border = THIN_BORDER
    for c in range(2, 8):
        cell = ws.cell(row=r, column=c)
        col_letter = get_column_letter(c)
        cell.value = f"=SUM({col_letter}2:{col_letter}{r-1})"
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.number_format = MONEY_FMT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _write_tva_sheet(ws, results):
    """Detail TVA breakdown per commune."""
    ws.sheet_properties.tabColor = "4CAF50"

    headers = ["Commune", "Type", "TVA 5,5% (HT)", "TVA 10% (HT)", "TVA 20% (HT)", "Total HT"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    for c in range(1, 7):
        ws.column_dimensions[get_column_letter(c)].width = 20

    r = 2
    for seg in sorted(results, key=lambda x: (-x.get("total_ht", 0))):
        synth = seg.get("synthesis", {})
        tva = synth.get("tva_groups", {})
        t55 = tva.get(0.055, {}).get("montant_ht", 0)
        t10 = tva.get(0.1, {}).get("montant_ht", 0)
        t20 = tva.get(0.2, {}).get("montant_ht", 0)
        total = t55 + t10 + t20

        row = [seg.get("commune", ""), seg.get("type", ""), t55, t10, t20, total]
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if c > 2:
                cell.number_format = MONEY_FMT
                cell.alignment = CENTER
        r += 1

    # Totals
    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=1).fill = TOTAL_FILL
    ws.cell(row=r, column=1).border = THIN_BORDER
    for c in range(3, 7):
        col_letter = get_column_letter(c)
        cell = ws.cell(row=r, column=c)
        cell.value = f"=SUM({col_letter}2:{col_letter}{r-1})"
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.number_format = MONEY_FMT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
