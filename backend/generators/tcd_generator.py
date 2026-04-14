"""
Generate tcd.xlsx (Tableau Croise Dynamique) for each commune/type segment.
This file serves as the structured data source for PDF generation.
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from backend.utils.logger import get_logger

log = get_logger()

# ── Styles ──
TITLE_FONT = Font(name="Montserrat", bold=True, size=13, color="1B5E20")
SECTION_FONT = Font(name="Montserrat", bold=True, size=10, color="2E7D32")
HEADER_FONT = Font(name="Montserrat", bold=True, size=9, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2E7D32")
DATA_FONT = Font(name="Montserrat", size=9)
DATA_BOLD = Font(name="Montserrat", bold=True, size=9)
TOTAL_FONT = Font(name="Montserrat", bold=True, size=10, color="1B5E20")
TOTAL_FILL = PatternFill("solid", fgColor="C8E6C9")
LABEL_FONT = Font(name="Montserrat", size=9, color="6D4C41")
VALUE_FONT = Font(name="Montserrat", bold=True, size=10)
MONEY_FMT = '#,##0.00 €'
PCT_FMT = '0.00%'
BORDER = Border(
    left=Side(style="thin", color="E0E0E0"),
    right=Side(style="thin", color="E0E0E0"),
    top=Side(style="thin", color="E0E0E0"),
    bottom=Side(style="thin", color="E0E0E0"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def generate_tcd(
    segment_df: pd.DataFrame,
    synthesis: dict,
    commune: str,
    work_type: str,
    output_path: Path,
) -> Path:
    """Generate tcd.xlsx for a given segment."""
    log.info(f"Generation TCD : {commune} / {work_type}")

    wb = Workbook()

    # Sheet 1: Synthese
    ws1 = wb.active
    ws1.title = "Synthese"
    ws1.sheet_properties.tabColor = "2E7D32"
    _write_synthese(ws1, synthesis, commune, work_type)

    # Sheet 2: Detail par Programme
    ws2 = wb.create_sheet("Detail Programme")
    ws2.sheet_properties.tabColor = "388E3C"
    _write_programmes(ws2, synthesis, segment_df)

    # Sheet 3: Detail par TVA
    ws3 = wb.create_sheet("Detail TVA")
    ws3.sheet_properties.tabColor = "4CAF50"
    _write_tva(ws3, synthesis)

    # Sheet 4: Detail par Installateur
    ws4 = wb.create_sheet("Detail Installateur")
    ws4.sheet_properties.tabColor = "66BB6A"
    _write_installateurs(ws4, segment_df)

    output_path.mkdir(parents=True, exist_ok=True)
    file_path = output_path / "tcd.xlsx"
    wb.save(file_path)
    return file_path


def _cell(ws, row, col, value, font=DATA_FONT, align=LEFT, fmt=None, fill=None):
    """Helper to write a styled cell."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.alignment = align
    c.border = BORDER
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    return c


def _write_synthese(ws, synthesis, commune, work_type):
    """Sheet 1: all key data for PDF generation."""
    type_labels = {
        "TEE": "Travaux d'Economie d'Energie",
        "PMR": "Travaux PMR (Accessibilite)",
        "INDIS_TEE": "Travaux Indissociablement Lies (TEE)",
        "INDIS_PMR": "Travaux Indissociablement Lies (PMR)",
    }

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 40

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    _cell(ws, r, 1, f"TABLEAU CROISE DYNAMIQUE - {commune.upper()}", TITLE_FONT)
    r += 2

    # Section: Identification
    _cell(ws, r, 1, "IDENTIFICATION", SECTION_FONT)
    r += 1
    fields = [
        ("Commune", commune),
        ("Type de travaux", type_labels.get(work_type, work_type)),
        ("Code type", work_type),
        ("Annee TFPB", "2024"),
        ("Annee des paiements", "2023"),
    ]
    for label, val in fields:
        _cell(ws, r, 1, label, LABEL_FONT)
        _cell(ws, r, 2, val, VALUE_FONT)
        r += 1

    r += 1

    # Section: Donnees fiscales
    _cell(ws, r, 1, "DONNEES FISCALES", SECTION_FONT)
    r += 1
    avis = "; ".join(str(a).strip() for a in synthesis.get("num_avis", []) if str(a).strip() not in ("", "nan"))
    fiscal = "; ".join(str(a).strip() for a in synthesis.get("num_fiscal", []) if str(a).strip() not in ("", "nan"))
    cdif = "; ".join(str(a).strip() for a in synthesis.get("cdif", []) if str(a).strip() not in ("", "nan"))
    adresse_cdif = "; ".join(str(a).strip()[:80] for a in synthesis.get("adresse_cdif", []) if str(a).strip() not in ("", "nan"))

    fiscal_fields = [
        ("N° d'avis d'imposition", avis or "N/A"),
        ("N° Fiscal", fiscal or "N/A"),
        ("CDIF / SIP", cdif or "N/A"),
        ("Adresse CDIF/SIP", adresse_cdif or "N/A"),
    ]
    for label, val in fiscal_fields:
        _cell(ws, r, 1, label, LABEL_FONT)
        _cell(ws, r, 2, val, DATA_BOLD, WRAP)
        r += 1

    r += 1

    # Section: Montants
    _cell(ws, r, 1, "MONTANTS", SECTION_FONT)
    r += 1
    montant_fields = [
        ("Nombre d'operations", synthesis.get("nb_operations", 0), None),
        ("Nombre de factures", synthesis.get("nb_factures", 0), None),
        ("Total HT Eligible", synthesis.get("total_ht_eligible", 0), MONEY_FMT),
        ("Total Subventions", synthesis.get("total_subventions", 0), MONEY_FMT),
        ("Base nette de degrevement", synthesis.get("base_nette", 0), MONEY_FMT),
        ("Taux de degrevement", 0.25, PCT_FMT),
        ("Montant de degrevement demande", synthesis.get("total_degrevement", 0), MONEY_FMT),
        ("Total virement TTC", synthesis.get("total_virement_ttc", 0), MONEY_FMT),
    ]
    for label, val, fmt in montant_fields:
        _cell(ws, r, 1, label, LABEL_FONT)
        c = _cell(ws, r, 2, val, VALUE_FONT, CENTER, fmt)
        r += 1

    # Highlight degrevement
    _cell(ws, r - 2, 1, "Montant de degrevement demande", TOTAL_FONT, fill=TOTAL_FILL)
    _cell(ws, r - 2, 2, synthesis.get("total_degrevement", 0), TOTAL_FONT, CENTER, MONEY_FMT, TOTAL_FILL)

    r += 1

    # Section: Programmes
    progs = synthesis.get("programmes", [])
    _cell(ws, r, 1, "PROGRAMMES CONCERNES", SECTION_FONT)
    r += 1
    _cell(ws, r, 1, "N° Programmes", LABEL_FONT)
    _cell(ws, r, 2, ", ".join(progs) if progs else "N/A", DATA_BOLD)


def _write_programmes(ws, synthesis, segment_df):
    """Sheet 2: detail by programme."""
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    _cell(ws, r, 1, "DETAIL PAR PROGRAMME", TITLE_FONT)
    r += 2

    headers = ["N° Programme", "Nb Factures", "HT Etudes", "HT Travaux", "HT Total"]
    for c, h in enumerate(headers, 1):
        _cell(ws, r, c, h, HEADER_FONT, CENTER, fill=HEADER_FILL)
    r += 1

    operations = synthesis.get("operations", {})
    if not operations:
        # Fallback: group by N° Programme from dataframe
        prog_col = "N° Programme"
        ht_col = "Montant des travaux éligibles retenus H.T"
        if prog_col in segment_df.columns:
            for prog, grp in segment_df.groupby(prog_col):
                prog_str = str(prog).strip()
                if prog_str in ("", "nan"):
                    continue
                ht = pd.to_numeric(grp[ht_col], errors="coerce").fillna(0).sum()
                _cell(ws, r, 1, prog_str, DATA_BOLD, CENTER)
                _cell(ws, r, 2, len(grp), DATA_FONT, CENTER)
                _cell(ws, r, 3, 0, DATA_FONT, CENTER, MONEY_FMT)
                _cell(ws, r, 4, float(ht), DATA_FONT, CENTER, MONEY_FMT)
                _cell(ws, r, 5, float(ht), DATA_BOLD, CENTER, MONEY_FMT)
                r += 1
    else:
        for op_id, op_info in sorted(operations.items()):
            _cell(ws, r, 1, op_id, DATA_BOLD, CENTER)
            _cell(ws, r, 2, op_info.get("nb_factures", 0), DATA_FONT, CENTER)
            _cell(ws, r, 3, op_info.get("montant_ht_etudes", 0), DATA_FONT, CENTER, MONEY_FMT)
            _cell(ws, r, 4, op_info.get("montant_ht_travaux", 0), DATA_FONT, CENTER, MONEY_FMT)
            _cell(ws, r, 5, op_info.get("montant_ht_total", 0), DATA_BOLD, CENTER, MONEY_FMT)
            r += 1

    # Total row
    _cell(ws, r, 1, "TOTAL", TOTAL_FONT, CENTER, fill=TOTAL_FILL)
    _cell(ws, r, 2, "", TOTAL_FONT, fill=TOTAL_FILL)
    for c in [3, 4, 5]:
        col_letter = get_column_letter(c)
        c_cell = _cell(ws, r, c, f"=SUM({col_letter}4:{col_letter}{r-1})", TOTAL_FONT, CENTER, MONEY_FMT, TOTAL_FILL)


def _write_tva(ws, synthesis):
    """Sheet 3: TVA breakdown."""
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    _cell(ws, r, 1, "DETAIL PAR TAUX DE TVA", TITLE_FONT)
    r += 2

    headers = ["Taux TVA", "Montant HT", "Nb Factures"]
    for c, h in enumerate(headers, 1):
        _cell(ws, r, c, h, HEADER_FONT, CENTER, fill=HEADER_FILL)
    r += 1

    tva_groups = synthesis.get("tva_groups", {})
    tva_labels = {0.055: "5,5%", 0.1: "10%", 0.2: "20%"}

    for rate in sorted(tva_groups.keys()):
        info = tva_groups[rate]
        label = tva_labels.get(rate, f"{rate*100:.1f}%")
        _cell(ws, r, 1, label, DATA_BOLD, CENTER)
        _cell(ws, r, 2, info.get("montant_ht", 0), DATA_FONT, CENTER, MONEY_FMT)
        _cell(ws, r, 3, info.get("nb_factures", 0), DATA_FONT, CENTER)
        r += 1

    # Total
    _cell(ws, r, 1, "TOTAL", TOTAL_FONT, CENTER, fill=TOTAL_FILL)
    col_b = get_column_letter(2)
    _cell(ws, r, 2, f"=SUM({col_b}4:{col_b}{r-1})", TOTAL_FONT, CENTER, MONEY_FMT, TOTAL_FILL)
    col_c = get_column_letter(3)
    _cell(ws, r, 3, f"=SUM({col_c}4:{col_c}{r-1})", TOTAL_FONT, CENTER, None, TOTAL_FILL)


def _write_installateurs(ws, segment_df):
    """Sheet 4: breakdown by installer."""
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    _cell(ws, r, 1, "DETAIL PAR INSTALLATEUR", TITLE_FONT)
    r += 2

    headers = ["Installateur", "Nb Factures", "Montant HT"]
    for c, h in enumerate(headers, 1):
        _cell(ws, r, c, h, HEADER_FONT, CENTER, fill=HEADER_FILL)
    r += 1

    inst_col = "Installateur"
    ht_col = "Montant des travaux éligibles retenus H.T"

    if inst_col not in segment_df.columns:
        return

    for inst, grp in segment_df.groupby(inst_col):
        inst_str = str(inst).strip()
        if inst_str in ("", "nan"):
            continue
        ht = pd.to_numeric(grp[ht_col], errors="coerce").fillna(0).sum()
        _cell(ws, r, 1, inst_str, DATA_FONT, LEFT)
        _cell(ws, r, 2, len(grp), DATA_FONT, CENTER)
        _cell(ws, r, 3, float(ht), DATA_FONT, CENTER, MONEY_FMT)
        r += 1

    # Total
    _cell(ws, r, 1, "TOTAL", TOTAL_FONT, LEFT, fill=TOTAL_FILL)
    col_b = get_column_letter(2)
    _cell(ws, r, 2, f"=SUM({col_b}4:{col_b}{r-1})", TOTAL_FONT, CENTER, None, TOTAL_FILL)
    col_c = get_column_letter(3)
    _cell(ws, r, 3, f"=SUM({col_c}4:{col_c}{r-1})", TOTAL_FONT, CENTER, MONEY_FMT, TOTAL_FILL)
